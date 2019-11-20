import openpyxl
from openpyxl import *
import os

Folder_path = os.getcwd()

# loop through files in directory and extract the customers' order number and the path of file


def listdir(directory):
	order_list = []
	order_list_location = []
	files = os.listdir(directory)
	# loop through every file in the directory
	
	for filename in files:
		# find files that are .msg types and format
		if filename.endswith('.msg'):
			internal_index_start = filename.lower().find('sob')
			if internal_index_start >= 0:
				index_end = internal_index_start + 8
				# find the string in the file name that matches our order number config
				order_number = filename[internal_index_start:index_end+1]
				# case that order number isn't entered correctly
				if order_number[3] == " ":
					order_number = order_number.replace(" ", "0")
					order_list.append(order_number)
					order_list_location.append(os.path.abspath(filename))
					# print("The External Sales Order No. is: " + order_number)
				else:
					order_list.append(order_number)
					order_list_location.append(os.path.abspath(filename))
					# print("The External Sales Order No. is: " + order_number)
		else:
			pass
	# return a list of tuples
	result = zip(order_list, order_list_location)
	return list(result)


dir_values = listdir(Folder_path)

wb = load_workbook(filename='Change log.xlsx')
# print(wb.sheetnames)
sheet = wb.active
	
for number in dir_values:
	link = number[1]
	currentRow = 2
	for value in sheet.iter_rows(min_row=2, values_only=True):
		if value[1] == number[0]:
			print(value[1])
			print(number[0])
			# loop through value in each row and create the cell as a hyperlink to the file directory
			sheet.cell(row=currentRow, column=2).hyperlink = link
			sheet.cell(row=currentRow, column=2).style = 'Hyperlink'
			currentRow += 1
			break
		else:
			currentRow += 1

# save to a new file
wb.save(filename='Change Log 1.xlsx')













