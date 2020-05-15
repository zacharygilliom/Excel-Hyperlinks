import openpyxl
from openpyxl import *
import os
from pathlib import Path
from orderNumbers import external_order_numbers
from tqdm import tqdm

# Proceeed with the understanding that we will be running the script in the current directory of all the emails
# and the change log.  The path can always be changed.
# ----------------------------------------------
# Windows WSL Path:
# /home/zach/python-projects/Excel-Hyperlinks/files/'
# ----------------------------------------------
# Linux Fedora Path:
# /home/zacharygilliom/Documents/python-projects/Excel-Hyperlinks/files/'

class emailDirectory:
    
    def __init__(self, directoryPath):
        self.directoryPath = directoryPath

    def listFiles(self):
        order_list = []
        files = os.listdir(self.directoryPath)
        # Loop through every file in the directory
        for file in files:
            # Find files that are .msg types and format
            if file.endswith('.msg') or file.endswith('.pdf'):
                order_list.append(file)
            else:
                pass
        return order_list


    def listFilePath(self):
        # Loop through the directory and append the paths to the file to a list.
        order_list_location = []
        files = os.listdir(self.directoryPath)
        # Path to network drive shared folder
        # pathToDir = '/home/zacharygilliom/Documents/python-projects/Excel-Hyperlinks/files/'
        pathToDir = '/home/zacharygilliom/python-projects/Excel-Hyperlinks/files/'
        for file in files:
            if file.endswith('.msg') or file.endswith('.pdf'):
                order_list_location.append(os.path.join(pathToDir, file))
        return order_list_location


    def zipFilesAndPath(self):
        # Create a tuple of our order list and order list location.
        file_and_path = zip(self.listFiles(), self.listFilePath())
        list_file_and_path = list(file_and_path)
        return list_file_and_path 



class emailMessage:
     
    def __init__(self, file, external_order_numbers):
        self.file = file
        self.external_order_numbers = external_order_numbers
       
    def getSplitMessage(self):
        # Create the base name of our file
        message_base = os.path.basename(self.file)
        # Split the base from the extension type (i.e. separate our the .msg, .pdf, etc...)
        split_base = os.path.splitext(message_base)
        # print(f'Split_base = {split_base}')
        message_no_ext = split_base[0].split()
        final_message = []
        for mess in message_no_ext:
            if len(mess) > 14:
               word = mess.split('-') 
               for w in word:
                   final_message.append(w)
            else:
                final_message.append(mess)
        return final_message


    def matchInternalOrder(self):
        order_list = self.getSplitMessage()
        # iterate with index through order list
        for i, word in enumerate(order_list): 
            word = str(word).casefold()
            if len(word) > 14:
                split_word = word.split('-')
                for j, words in enumerate(split_word):
                    words = str(words).casefold()
                    if words[:3] == 'ajh':
                        if len(words) == 9 and words[3] == '0':
                            word = words[:3] + '0' + words[4:]
                            return word
                        elif len(words) == 9 and words[3] != '0':
                            word = words[:3] + '0' + words[4:]
                            return word
                        elif len(words) == 8: 
                            word = words[:3] + '0' + words[3:]
                            return word
                        elif len(words) == 3:
                            if len(words[j+1]) == 6:
                                word = split_word[j] + split_word[j+1]
                                return word
                            elif len(words[j+1]) == 5:
                                word = split_word[j] + '0' + split_word[j+1]
                                return word
            else:
            # casefold is a string method to ignore case
            # if one of the words in the list is equal to ajh, below are all the different variants of how it could be.
                if word[:3] == 'ajh':
                    if len(word) == 9 and word[3] == '0':
                        return word[:3] + '0' + word[4:]
                    elif len(word) == 9 and word[3] != '0':
                        return word[:3] + '0' + word[4:]
                    elif len(word) == 8: 
                        return word[:3] + '0' + word[3:]
                    elif len(word) == 3:
                        if len(order_list[i+1]) == 6:
                            new_word = order_list[i] + order_list[i+1]
                            return new_word
                        elif len(order_list[i+1]) == 5:
                            new_word = order_list[i] + '0' + order_list[i+1]
                            return new_word

    def matchExternalOrder(self):
        ext_order_numbers = self.external_order_numbers
        # External order numbers can have two different lengths, 4 or 5, but they are generally fully type out and on
        # differently typed like they are in internal order numbers.
        for word in self.getSplitMessage():
            word = str(word).casefold()
            for number in ext_order_numbers:
                if word[:5] == number.casefold():
                    return word
                elif word[:4] == number.casefold():
                    return word

def linkFiles(workbook, direc, ext_order_numbers):
    # Activate the sheet in our workbook.
    for sheet in workbook.worksheets:
        # sheet = workbook.active
        # Loop through the directory and create an instance of the emailMessage class.
        for dir_value in tqdm(direc):
            # [0] in dir_value is our order number, [1] is our file location of the file.
            link_number = emailMessage(dir_value[0], ext_order_numbers)
            link_location = dir_value[1]
            if link_number.matchInternalOrder():
                linked_val = link_number.matchInternalOrder()
            elif link_number.matchExternalOrder():
                linked_val = link_number.matchExternalOrder()
            else:
                linked_val = ""
            currentRow = 2
            for value in sheet.iter_rows(min_row=2, values_only=True):
                if sheet.cell(row=currentRow, column=2).hyperlink is None:
                    if value[1].casefold() == linked_val.casefold():
                        sheet.cell(row=currentRow, column=2).hyperlink = link_location
                        sheet.cell(row=currentRow, column=2).style = 'Hyperlink'
                        currentRow += 1
                    else:
                        currentRow += 1
                else:
                    currentRow += 1
                    
        # WSL 2 location
        # workbook.save(filename='/home/zach/python-projects/Excel-Hyperlinks/workbooks/Change Log Updated.xlsx')
        # Fedora Desktop Linux location
        # workbook.save(filename='/home/zacharygilliom/Documents/python-projects/Excel-Hyperlinks/workbooks/Change Log Updated.xlsx')
        # Fedora Laptop Linux location
        workbook.save(filename='/home/zacharygilliom/python-projects/Excel-Hyperlinks/workbooks/Change Log Updated.xlsx')

def main():

    # Specify the path to the source of our email files to link to our workbook.
    # Fedora Desktop Linux path
    # Folder_path = Path("/home/zacharygilliom/Documents/python-projects/Excel-Hyperlinks/files/")
    # WSL 2 path
    # Folder_path = Path("/home/zach/python-projects/Excel-Hyperlinks/files/")
    # Fedora Laptop Linux Path
    Folder_path = Path('/home/zacharygilliom/python-projects/Excel-Hyperlinks/files/')

    # Speciy the external order numbers that we will match.
    ext_order_numbers = external_order_numbers

    # Create an instance of our emailDirectory class.
    userDirectory = emailDirectory(Folder_path)

    # The zipFilesAndPath method will createa a tuple of the order numbers and the path to the file.
    userDirectoryFiles = userDirectory.zipFilesAndPath()

    # Open up our workbook specifying the path to it via openpxl method.
    # Fedora Desktop Linux path
    # book = load_workbook(filename='/home/zacharygilliom/Documents/python-projects/Excel-Hyperlinks/workbooks/Change Log.xlsx')
    # WSL 2 path
    # book = load_workbook(filename='/home/zach/python-projects/Excel-Hyperlinks/workbooks/Change Log.xlsx')
    # Fedora Laptop Linux path
    book = load_workbook(filename='/home/zacharygilliom/python-projects/Excel-Hyperlinks/workbooks/Change Log.xlsx')

    # call function with our open workbook and our directory class.
    linkFiles(workbook=book, direc = userDirectoryFiles, ext_order_numbers=ext_order_numbers)


if __name__ == "__main__":
    main()

