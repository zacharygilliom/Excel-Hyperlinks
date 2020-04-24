import openpyxl
from openpyxl import *
import os
from pathlib import Path

# Proceeed with the understanding that we will be running the script in the current directory of all the emails
# and the change log.  The path can always be changed.

# Linux Path
Folder_path = Path("/home/zach/python-projects/Excel-Hyperlinks/files/")

# Windows Path
# Folder_path = Path("C:\\Users\\zacha\\Documents\\python-projects\\Excel-Hyperlinks\\files")


ext_order_numbers = ['KLJH', 'AJHYN', 'OPJD']


# This function helps deal with the different customers' orders.
# Different customers have different alphabetical notations
def match_external_order(message, external_numbers):
    base = os.path.basename(message)
    no_ext_message = os.path.splitext(base)
    split_message = no_ext_message[0].split()
    for word in split_message:
        for number in external_numbers:
            if word[:5] == number:
                return word
        # if word[:5] in external_numbers:
            # return word
            elif word[:4] == number:
                return word
            else:
                pass


# for internal order numbers, the main thing we need to worry about is the dropping of the "0"
def match_internal_order(message):
    base = os.path.basename(message)
    no_ext_message = os.path.splitext(base)
    split_message = no_ext_message[0].split()
    for word in split_message:
        if word[:3].lower() == 'ajh' and len(word) < 11:
            if len(word) == 9:
                return word
            else:
                new_word = word[:3] + "0" + word[3:]
                return new_word
        else:
            pass


# loop through files in directory and extract the customers' order number and the path of file
def listdir(directory):
    order_list = []
    order_list_location = []
    files = os.listdir(directory)
    # loop through every file in the directory
    for filename in files:
        # find files that are .msg types and format
        if filename.endswith('.msg'):
            # go through each file and match it to either the internal order number or external order number
            if match_external_order(filename, ext_order_numbers):
                order_number = match_external_order(filename, ext_order_numbers)
                order_list.append(order_number)
                order_list_location.append(os.path.abspath(f'files/{filename}'))
            elif match_internal_order(filename):
                order_number = match_internal_order(filename)
                order_list.append(order_number)
                order_list_location.append(os.path.abspath(f'files/{filename}'))
            else:
                pass
    # return a list of pairs.  We want both the order number and the location for the link to work in excel
    print(order_list[0])
    print(order_list_location[0])
    result = zip(order_list, order_list_location)
    return list(result)


dir_values = listdir(Folder_path)
# print(dir_values)

wb = load_workbook(filename='Change Log.xlsx')
sheet = wb.active
# now we have our excel book open we want to loop through every number in our order number column and match it to
# one of the values from our directory list
for number in dir_values:
    link = number[1]
    currentRow = 2
    for value in sheet.iter_rows(min_row=2, values_only=True):
        if value[1] == number[0]:
            # loop through value in each row and create the cell as a hyperlink to the file directory
            sheet.cell(row=currentRow, column=2).hyperlink = link
            sheet.cell(row=currentRow, column=2).style = 'Hyperlink'
            currentRow += 1
            break
        else:
            currentRow += 1

# save to a new file
wb.save(filename='Change Log Updated.xlsx')
